from io import BytesIO

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.http import HttpResponseNotFound
from django.urls import reverse, reverse_lazy
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from openpyxl import load_workbook

from base.forms import CuoteroForm, ComisionMoraForm, VendedorForm, ClienteForm, UsuarioForm, TipoDocumentoForm, \
    ImportadorClienteForm, UsuarioChangeForm
from base.mixins import AdminMixin
from base.models import Cuotero, ComisionMora, Vendedor, Cliente, TipoDocumento


class SiteLoginView(LoginView):
    template_name = 'login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        user = self.request.user
        if not user.is_authenticated:
            return HttpResponseNotFound('<h1>Por favor ingrese al sistema</h1>')
        else:
            if user.is_superuser:
                return reverse('admin-home')  # TODO vista de superuser
            elif user.groups.filter(name='Admin').exists():
                return reverse('admin-home')
            elif user.groups.filter(name='Agente').exists():
                return reverse('home')
        logout(self.request)
        messages.add_message(self.request, messages.WARNING, 'Usuario inválido.')
        return reverse('login')


class HomeView(TemplateView):
    template_name = 'home.html'

    def dispatch(self, request, *args, **kwargs):
        if not self.request.user.is_authenticated:
            raise PermissionDenied("403")
        return super(HomeView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(HomeView, self).get_context_data(**kwargs)
        context["user"] = self.request.user
        return context


class DashboardView(TemplateView, AdminMixin):
    template_name = 'dashboard.html'


# CUOTERO
class CuoteroListView(ListView, AdminMixin):
    template_name = 'cuotero/list.html'
    model = Cuotero
    context_object_name = "cuoteros"


class CuoteroDetailView(DetailView, AdminMixin):
    template_name = 'cuotero/detail.html'
    model = Cuotero
    context_object_name = 'cuotero'
    pk_url_kwarg = 'cuotero_id'


class CuoteroCreateView(CreateView, AdminMixin):
    template_name = 'cuotero/create.html'
    model = Cuotero
    success_url = reverse_lazy('cuotero.list')
    form_class = CuoteroForm


class CuoteroUpdateView(UpdateView, AdminMixin):
    template_name = 'cuotero/update.html'
    model = Cuotero
    context_object_name = 'cuotero'
    pk_url_kwarg = 'cuotero_id'
    form_class = CuoteroForm

    def get_success_url(self):
        return reverse('cuotero.detail', kwargs={'cuotero_id': self.object.id})


class CuoteroDeleteView(DeleteView, AdminMixin):
    template_name = 'cuotero/delete.html'
    model = Cuotero
    context_object_name = 'cuotero'
    pk_url_kwarg = 'cuotero_id'
    success_url = reverse_lazy('cuotero.list')


# COMISION Y MORA
class ComisionMoraListView(ListView, AdminMixin):
    template_name = 'comisionmora/list.html'
    model = ComisionMora
    context_object_name = "comisiones_moras"


class ComisionMoraDetailView(DetailView, AdminMixin):
    template_name = 'comisionmora/detail.html'
    model = ComisionMora
    context_object_name = 'comision_mora'
    pk_url_kwarg = 'comision_mora_id'


class ComisionMoraCreateView(CreateView, AdminMixin):
    template_name = 'comisionmora/create.html'
    model = ComisionMora
    success_url = reverse_lazy('comision_mora.list')
    form_class = ComisionMoraForm


class ComisionMoraUpdateView(UpdateView, AdminMixin):
    template_name = 'comisionmora/update.html'
    model = ComisionMora
    context_object_name = 'comision_mora'
    pk_url_kwarg = 'comision_mora_id'
    form_class = ComisionMoraForm

    def get_success_url(self):
        return reverse('comision_mora.detail', kwargs={'comision_mora_id': self.object.id})


class ComisionMoraDeleteView(DeleteView, AdminMixin):
    template_name = 'comisionmora/delete.html'
    model = ComisionMora
    context_object_name = 'comision_mora'
    pk_url_kwarg = 'comision_mora_id'
    success_url = reverse_lazy('comision_mora.list')


# VENDEDOR
class VendedorListView(ListView, AdminMixin):
    template_name = 'vendedor/list.html'
    model = Vendedor
    context_object_name = "vendedores"


class VendedorDetailView(DetailView, AdminMixin):
    template_name = 'vendedor/detail.html'
    model = Vendedor
    context_object_name = 'vendedor'
    pk_url_kwarg = 'vendedor_id'


class VendedorCreateView(CreateView, AdminMixin):
    template_name = 'vendedor/create.html'
    model = Vendedor
    success_url = reverse_lazy('vendedor.list')
    form_class = VendedorForm


class VendedorUpdateView(UpdateView, AdminMixin):
    template_name = 'vendedor/update.html'
    model = Vendedor
    context_object_name = 'vendedor'
    pk_url_kwarg = 'vendedor_id'
    form_class = VendedorForm

    def get_success_url(self):
        return reverse('vendedor.detail', kwargs={'vendedor_id': self.object.id})


class VendedorDeleteView(DeleteView, AdminMixin):
    template_name = 'vendedor/delete.html'
    model = Vendedor
    context_object_name = 'vendedor'
    pk_url_kwarg = 'vendedor_id'
    success_url = reverse_lazy('vendedor.list')


# CLIENTE
class ClienteListView(LoginRequiredMixin, ListView):
    template_name = 'cliente/list.html'
    model = Cliente
    context_object_name = "clientes"


class ClienteDetailView(LoginRequiredMixin, DetailView):
    template_name = 'cliente/detail.html'
    model = Cliente
    context_object_name = 'cliente'
    pk_url_kwarg = 'cliente_id'


class ClienteCreateView(LoginRequiredMixin, CreateView):
    template_name = 'cliente/create.html'
    model = Cliente
    success_url = reverse_lazy('cliente.list')
    form_class = ClienteForm


class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'cliente/update.html'
    model = Cliente
    context_object_name = 'cliente'
    pk_url_kwarg = 'cliente_id'
    form_class = ClienteForm

    def get_success_url(self):
        return reverse('cliente.detail', kwargs={'cliente_id': self.object.id})


class ClienteDeleteView(DeleteView, AdminMixin):
    template_name = 'cliente/delete.html'
    model = Cliente
    context_object_name = 'cliente'
    pk_url_kwarg = 'cliente_id'
    success_url = reverse_lazy('cliente.list')


# USUARIO
class UsuarioListView(ListView, AdminMixin):
    template_name = 'usuario/list.html'
    model = User
    context_object_name = "usuarios"


class UsuarioDetailView(DetailView, AdminMixin):
    template_name = 'usuario/detail.html'
    model = User
    context_object_name = 'usuario'
    pk_url_kwarg = 'usuario_id'


class UsuarioCreateView(CreateView, AdminMixin):
    template_name = 'usuario/create.html'
    model = User
    success_url = reverse_lazy('usuario.list')
    form_class = UsuarioForm


class UsuarioUpdateView(UpdateView, AdminMixin):
    template_name = 'usuario/update.html'
    model = User
    context_object_name = 'usuario'
    pk_url_kwarg = 'usuario_id'
    success_url = reverse_lazy('usuario.list')
    form_class = UsuarioChangeForm


class UsuarioDeleteView(DeleteView, AdminMixin):
    template_name = 'usuario/delete.html'
    model = User
    context_object_name = 'usuario'
    pk_url_kwarg = 'usuario_id'
    success_url = reverse_lazy('usuario.list')


# TIPO DOCUMENTO
class TipoDocumentoListView(ListView, AdminMixin):
    template_name = 'tipo_documento/list.html'
    model = TipoDocumento
    context_object_name = "tipos_documento"


class TipoDocumentoDetailView(DetailView, AdminMixin):
    template_name = 'tipo_documento/detail.html'
    model = TipoDocumento
    context_object_name = 'tipo_documento'
    pk_url_kwarg = 'tipo_documento_id'


class TipoDocumentoCreateView(CreateView, AdminMixin):
    template_name = 'tipo_documento/create.html'
    model = TipoDocumento
    success_url = reverse_lazy('tipo_documento.list')
    form_class = TipoDocumentoForm


class TipoDocumentoUpdateView(UpdateView, AdminMixin):
    template_name = 'tipo_documento/update.html'
    model = TipoDocumento
    context_object_name = 'tipo_documento'
    pk_url_kwarg = 'tipo_documento_id'
    form_class = TipoDocumentoForm

    def get_success_url(self):
        return reverse('tipo_documento.detail', kwargs={'tipo_documento_id': self.object.id})


class TipoDocumentoDeleteView(DeleteView, AdminMixin):
    template_name = 'tipo_documento/delete.html'
    model = TipoDocumento
    context_object_name = 'tipo_documento'
    pk_url_kwarg = 'tipo_documento_id'
    success_url = reverse_lazy('tipo_documento.list')


class ImportadorClienteView(FormView):
    template_name = "cliente/importador.html"
    form_class = ImportadorClienteForm

    def procesar_xls(self, file):
        response = {
            "success": True,
            "message": "Importados los datos con éxito."
        }
        wb = load_workbook(filename=BytesIO(file.read()))
        ws = wb.worksheets[0]
        rowCount = 0
        errores = []
        tipo_documento = TipoDocumento.objects.filter(nombre__icontains="CEDULA").first()
        tipo_documento = tipo_documento if tipo_documento else TipoDocumento.objects.filter(nombre__icontains="CÉDULA").first()
        try:
            with transaction.atomic():
                for row in ws.iter_rows():
                    rowCount += 1
                    nombre = row[0].value
                    documento = str(row[1].value)
                    direccion = row[2].value
                    cliente, created = Cliente.objects.update_or_create(
                        nombre=nombre,
                        tipo_documento=tipo_documento if tipo_documento else None,
                        ci=documento,
                        direccion=direccion,
                    )
        except Exception as e:
            errores.append(rowCount)

        if len(errores):
            response['success'] = False
            response['message'] = "Errores en las filas: {}".format(', '.join(str(e) for e in errores))

        return response

    def post(self, request, *args, **kwargs):
        response = {}
        if request.FILES.get('documento', False):
            file = request.FILES['documento']
            filename = file.name
            extension = filename.split('.')[-1]
            if extension in ['xls', 'xlsx']:
                response = self.procesar_xls(file)
            else:
                response = {
                    "success": False,
                    "message": "Formato de archivo no válido."
                }
            return self.render_to_response(self.get_context_data(**response))
        return self.render_to_response(self.get_context_data())

